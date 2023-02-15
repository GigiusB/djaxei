from collections import OrderedDict

from django.apps import apps
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from djaxei.exceptions import ImportException


class Importer:
    def __init__(self, modems: list):
        self.modems = modems

    def xls_import(self, file):
        wb = load_workbook(file, read_only=True, data_only=True)

        remappings = {
            sheet_name: {} for sheet_name in wb.sheetnames
        }

        for modem in self.modems:
            loader = modem.loader(remappings, modem.model_label, wb.worksheets[wb.sheetnames.index(modem.model_label)])
            loader()
        print(wb)


class LoaderFx:
    def __init__(self, model_name: str, modifier=None):
        self.model_name = model_name
        self.modifier = modifier
        self.klass = apps.get_model(*model_name.split('.'))

    def __call__(self, mappings: dict, worksheet: Worksheet):
        for counter, row in enumerate(worksheet.rows):
            if counter == 0:  # header
                headers = [c.value for c in row][1:]
            else:
                oldid = row[0].value
                data = {h: row[z + 1].value for z, h in enumerate(headers)}

                if self.modifier:
                    self.modifier(data, mappings)

                new_obj = self.klass.objects.create(**data)
                mappings[self.model_name][oldid] = new_obj.id



class Importer_:
    def __init__(self, tmpdir=None, **kwargs):
        self.tmpdir = tmpdir

    def xls_import(self, file, models_dict, *args, **kwargs):
        wb = load_workbook(file, data_only=True)

        for ws_name, callback in models_dict.items():
            ws = wb[ws_name]
            for rownum, row in enumerate(ws.iter_rows()):
                row = [cell.value for cell in row]
                if rownum == 0:
                    header = row
                else:
                    try:
                        data_dict = OrderedDict(zip(header, row))
                        data_dict = {k: v for k, v in data_dict.items() if k != None}
                        callback(data_dict)
                    except Exception as e:
                        raise ImportException(e, worksheet=ws_name)
        wb.close()

