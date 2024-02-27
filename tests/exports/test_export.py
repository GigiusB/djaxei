# see https://github.com/pytest-dev/pytest-django/blob/master/pytest_django/fixtures.py
import datetime
import json
import random
from collections.abc import Iterable
from tempfile import NamedTemporaryFile

import pytest
from demoproject.app1.models import (DemoModel1, DemoModel2,
                                     DemoModel3, VeryLongNameModelDemoModel4,)
from django.db.models import Q, QuerySet
from openpyxl.reader.excel import load_workbook

from djaxei import Exporter
from djaxei.modems.field import DatetimeNonAwareModem, JsonToStringModem
from djaxei.modems.model import FieldListModelMoDem


def root_generator(key):
    qs = DemoModel1.objects.all()
    return {
        1: random.choice(qs),
        2: qs.filter(id__in=random.sample(list(qs.values_list('id', flat=True)), 2)),
        3: random.sample(list(qs), 2)
    }[key]



@pytest.mark.parametrize(
    'root_fx_key, m1, m2, m3, m4',
    (
        pytest.param(
            1, 'app1.demomodel1', 'app1.demomodel2', 'app1.DemoModel3', 'app1.VeryLongNameModelDemoModel4',
            id='strings'
        ),
        pytest.param(
            2, DemoModel1, DemoModel2, DemoModel3, VeryLongNameModelDemoModel4,
            id='models'
        ),
        pytest.param(
            3, DemoModel1, 'app1.demomodel2', DemoModel3, VeryLongNameModelDemoModel4,
            id='mixed'
        ),
    )
)
def test_exporter(root_fx_key, m1, m2, m3, m4, recordset):
    roots = root_generator(root_fx_key)

    fx_dt = lambda dt: dt.replace(tzinfo=None)
    modems = {
        FieldListModelMoDem(
            model=m1,
            fields=['id', 'fk_id', 'char', 'integer', 'logic', 'null_logic',
                    'date', 'nullable', 'choice',
                    DatetimeNonAwareModem('timestamp'), JsonToStringModem('j')]
        ),
        FieldListModelMoDem(
            m2,
            ['id', 'fk_id', 'integer']
        ),
        FieldListModelMoDem(
            m3,
            ['id', 'fk_id', 'char', 'integer']
        ),
        FieldListModelMoDem(
            m4,
            ['id', 'fk3_id', 'char', 'fk2_id', 'integer']
        ),
    }

    exporter = Exporter(root=roots, modems=modems)
    with NamedTemporaryFile(suffix='.xlsx') as fo:
        exporter.xls_export(fo)

        # Checks starts here

        # Read generated workbook
        results = {}
        wb = load_workbook(filename=fo.name, read_only=True)
        assert sorted(wb.sheetnames) == ['app1.demomodel1', 'app1.demomodel2', 'app1.demomodel3', 'app1.verylongnamemodeldemomodel4']

        for sheet in wb.worksheets:
            results[sheet.title] = []
            for row in sheet.rows:
                results[sheet.title].append([c.value for c in row])

    if isinstance(roots, QuerySet):
        roots = list(roots.all())
    elif not isinstance(roots, Iterable):
        roots = [roots]
    root_class = roots[0].__class__
    root_model_name = root_class._meta.label_lower

    result_root_header = results[root_model_name][0]

    # check roots
    expected = list(root_class.objects.filter(id__in=[c.id for c in roots]).values())
    for e in expected:
        e['j'] = json.dumps(e['j'])
        e['timestamp'] = fx_dt(e['timestamp']).replace(microsecond=0)
        e['date'] = datetime.datetime.fromordinal(e['date'].toordinal())
    actual = [dict(zip(result_root_header, row)) for row in results[root_model_name][1:]]
    for k in actual:
        k['timestamp'] = k['timestamp'].replace(microsecond=0)
    assert sorted(expected, key=lambda x: x['id']) == sorted(actual, key=lambda x: x['id'])

    root_ids = [x.id for x in roots]

    # check DemoModel2
    ids2 = sorted([x[0] for x in results['app1.demomodel2'][1:]])
    assert sorted(list(DemoModel2.objects.filter(fk_id__in=root_ids).values_list('id', flat=True))) == ids2

    # check DemoModel3
    ids3 = sorted([x[0] for x in results['app1.demomodel3'][1:]])
    assert sorted(list(DemoModel3.objects.filter(fk_id__in=root_ids).values_list('id', flat=True))) == ids3

    # check VeryLongNameModelDemoModel4
    ids4 = sorted([x[0] for x in results['app1.verylongnamemodeldemomodel4'][1:]])
    assert sorted(
        list(VeryLongNameModelDemoModel4.objects.filter(Q(fk3__fk__in=root_ids) | Q(fk2__fk__in=root_ids)).values_list('id', flat=True))
    ) == ids4
